from enum import unique
from itertools import count
from pipes import Template
import re
import streamlit as st
import openpyxl
import pandas as pd
import os
import json

st.set_page_config(layout="wide")
st.header('Excel Tools')
excel_file_path = '/Volumes/research/Downloads-Other/data1.xlsx'
export_path = '/Volumes/research/Downloads-Other/csv-exports/'
export_temp_path = '/Volumes/research/Downloads-Other/csv-exports/temp/'






def consolidate_CSVs_to_one_file(csv_file_path_list, column_map):
    if os.path.exists(f'{export_path}column_map.json'): os.remove(f'{export_path}column_map.json')
    # Create an empty DataFrame for the consolidated data
    consolidated_df = pd.DataFrame()

    # Iterate through the list of CSV files
    for file_path in csv_file_path_list:
        print(f"beginning {file_path}")
        # Read the current CSV file
        if '17' in file_path:
            print('here')
        df = pd.read_csv(f"{export_path}{file_path}")
        start_col_list = df.columns.tolist()
        # Rename columns based on the column_map
        df.rename(columns=column_map, inplace=True)

        # Select only the columns that are in the column_map
        df = df[[col for col in column_map.values() if col in df.columns]]

        # Reset the index of df to avoid index conflicts
        df.reset_index(drop=True, inplace=True)
        end_col_list = df.columns.tolist()

        # Add the 'Entity' column with the file_path as its value for each row
        df['Entity'] = file_path

        with open(f'{export_path}all_csv_headers.json', 'a') as fp:
            fp.write(f"{file_path}\n")
            fp.write(f"{start_col_list}\n")
            fp.write(f"{end_col_list}\n\n\n")

        # Append the DataFrame to the consolidated DataFrame
        consolidated_df = pd.concat([consolidated_df, df], ignore_index=True)

        print(f"completed {file_path}")

    # Export the consolidated DataFrame to a new CSV file
    consolidated_df.to_csv('model-json/consolidated.csv', index=False)

    print('CSV files have been consolidated into consolidated.csv')




def get_file_paths_from_files_in_folder(folder_path, file_extension_list=[]):
    
    # Get all files in the folder
    files_in_directory = os.listdir(folder_path)

    #check if any file extensions to filter were given
    if len(file_extension_list) >0:
        #filter that list to only files that have the extension that we're looking for  
        filtered_list = [file for file in files_in_directory 
                        if any(file.endswith(ext) for ext in file_extension_list)]
        return filtered_list
    else:
        # if no file extensions were given, return the full list of files
        return files_in_directory 

def get_first_lines_of_file(file_path, nlines=3):
    with open(file_path) as f:
        first_lines = [next(f) for x in range(nlines)]
    return first_lines


def get_first_rows_of_csv(csv_file_path, nrows=3):
    df = pd.read_csv(csv_file_path, nrows=3)
    return df

def cleanse_sheet_name(sheet_name):
    # Replace forbidden characters with an underscore or another suitable character
    cleansed_name = re.sub(r'[\\/*?:"<>|]', '_', sheet_name)

    # Remove leading and trailing spaces
    cleansed_name = cleansed_name.strip()

    # Optionally, truncate the name if it's too long
    max_length = 255  # Maximum filename length. Adjust as needed.
    if len(cleansed_name) > max_length:
        cleansed_name = cleansed_name[:max_length]

    return cleansed_name

def inspect_excel_file(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet_inspection_list = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_inspection_list.append(
            {
                'sheet_name': sheet,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
            }
        )

def get_title_row_number(csv_file_path, title_row_identifer='Id', within_n_rows=3):
    with open(csv_file_path) as f:
        print(csv_file_path)
        all_data = f.read()
        f.seek(0)
        number_of_rows = all_data.count('\n')
        rannge_val = min(number_of_rows-1, within_n_rows)
        first_lines = [next(f) for x in range(rannge_val)]
    count = 0
    for line in first_lines:
        count += 1
        if title_row_identifer in line:
            return count

def export_excel_sheet_to_csv(excel_file_path, sheet_name, export_path='temp/', column_range='A:ZZ', row_range=1100, title_row = None ):
    # Read in the data for the desired sheet
    # Specify the sheet name and the range you want to read
    # For example, 'A1:D10' reads cells from A1 to D10
    try:
        if title_row == None:
            df = pd.read_excel(excel_file_path, 
                        sheet_name=sheet_name)
                        # , 
                        # usecols=column_range,
                        # nrows=row_range)
        else:
            df = pd.read_excel(excel_file_path, 
                        sheet_name=sheet_name, 
                        header=title_row)
        # Export to CSV
        # You can specify the file name for your CSV
        
        export_file_path = f"{export_path}{cleanse_sheet_name(sheet_name)}.csv"
        if os.path.exists(export_file_path): os.remove(export_file_path)
        result = df.to_csv(export_file_path, index=False, mode='w')
        return export_file_path
    except Exception as e:
        print(e)
        export_file_path = None
        return False

def find_header_row_number(csv_file_path, title_row_identifer='Id', within_n_rows=10):
    with open(csv_file_path) as f:
        all_data = f.read()
        f.seek(0)
        number_of_rows = all_data.count('\n')
        rannge_val = min(number_of_rows, within_n_rows)
        first_lines = [next(f) for x in range(rannge_val)]
    count = 0
    for line in first_lines:
        count += 1
        if title_row_identifer in line:
            return count-1


def export_all_sheets_to_csv(excel_file_path, export_path, column_range='A:ZZ', row_range=1100, title_row_identifer='Id'):    
    # Read the Excel file
    # Get the sheet names
    sheets_name_list = pd.ExcelFile(excel_file_path).sheet_names
    # Iterate over each sheet name
    CSV_Files = []
    for sheet_name in sheets_name_list:
        # Read in the data for the desired sheet and export raw data to CSV
        print(sheet_name)
        raw_csv = export_excel_sheet_to_csv(excel_file_path, sheet_name, export_path=export_temp_path, column_range='A:ZZ', row_range=1100)
        if not raw_csv:
            print(f"Could not export {sheet_name} to CSV")
            continue
        # Find the row number of the header row in the raw csv
        title_row_number = find_header_row_number(raw_csv, title_row_identifer=title_row_identifer, within_n_rows=5)
        #Re-export the sheet to CSV with the header row number
        csv_with_headers = export_excel_sheet_to_csv(excel_file_path, sheet_name, export_path, column_range='A:ZZ', row_range=1100, title_row=title_row_number)
        CSV_Files.append(csv_with_headers)
    return CSV_Files

def inspect_excel_file(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet_inspection_list = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_inspection_list.append(
            {
                'sheet_name': sheet,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
            }
        )
        df = pd.DataFrame(sheet_inspection_list)
    return df

def list_excel_sheets(excel_file_path):
    # Get the sheet names as a list
    sheets_list = pd.ExcelFile(excel_file_path).sheet_names
    # Convert the list to a Pandas DataFrame
    df = pd.DataFrame(sheets_list, columns=['Worksheet Names'])
    return df

def extract_field_data(df, field, rows):
    if field in df.columns:
        # Get the first 10 values of the field, converted to string
        value_list =  '|'.join(df[field].astype(str).head(rows).tolist())
        unique_value_list = list(set(value_list.split('|')))
        return unique_value_list
    else:
        return None

def get_distinct_list_of_titles(export_path):
    csv_file_paths = get_file_paths_from_files_in_folder(export_path, ['.csv'])
    all_titles = set()
    for file_path in csv_file_paths:
        full_file_path = f"{export_path}{file_path}"
        first_3_lines = get_first_lines_of_file(full_file_path, 3)
        count = 0
        for line in first_3_lines:
            count += 1
            if 'Name,Type' in line:
                titles = line.split(',')
                for title in titles:
                    all_titles.add(title)
    distinct_titles = list(set(all_titles))
    sorted_titles = sorted(distinct_titles)
    return sorted_titles

def get_dict__of_dataframes_of_all_csvs_in_folder(export_path):
    csv_file_paths = get_file_paths_from_files_in_folder(export_path, ['.csv']) 
    dataframes_dict = {}
    for file_path in csv_file_paths:
        full_file_path = f"{export_path}{file_path}"
        df = pd.read_csv(full_file_path)
        dataframes_dict[file_path] = df
    return dataframes_dict    
    

if st.button("List Excel Sheets"):
        st.write(list_excel_sheets(excel_file_path))

if st.button("Inspect Excel File"):
        st.write(inspect_excel_file(excel_file_path))

if st.button('Export Sheets to CSV'):
    export_all_sheets_to_csv(excel_file_path, export_path, column_range='A:ZZ', row_range=1100, title_row_identifer='Name,Type')

if st.button('List CSV Files in Folder'):
    file_extension_list = ['.csv']
    st.write(get_file_paths_from_files_in_folder(export_path, file_extension_list))

if st.button('Get the First 3 Rows of all CSV Files in Folder'):
    csv_file_paths = get_file_paths_from_files_in_folder(export_path, ['.csv'])
    for file_path in csv_file_paths:
        full_file_path = f"{export_path}{file_path}"
        st.write(get_first_rows_of_csv(full_file_path, 3))

if st.button('Find the column title row in all CSV Files in Folder'):
    csv_file_paths = get_file_paths_from_files_in_folder(export_path, ['.csv'])
    for file_path in csv_file_paths:
        full_file_path = f"{export_path}{file_path}"
        st.subheader(file_path, divider="red")
        first_3_lines = get_first_lines_of_file(full_file_path, 3)
        count = 0
        for line in first_3_lines:
            count += 1
            if 'Name,Type' in line:
                st.write(count)
                # st.write(line)

if st.button('Find all title rows in all CSV Files in Folder'):
    csv_file_paths = get_file_paths_from_files_in_folder(export_path, ['.csv'])
    list_of_title_rows = []
    set_of_title_rows = set()
    for file_path in csv_file_paths:
        full_file_path = f"{export_path}{file_path}"
        title_row_number = get_title_row_number(full_file_path, title_row_identifer='Name,Type', within_n_rows=3)
        list_of_title_rows.append(title_row_number)
        set_of_title_rows.add(title_row_number)
    st.write(list_of_title_rows)
    st.write(set_of_title_rows)


if st.button('Create the Column Title Normalization Template'):
    sorted_titles = get_distinct_list_of_titles(export_path)
    Template = {}
    for title in sorted_titles:
        Template[title] = title
    
    template_string = json.dumps(Template, indent=4)
    st.write(Template)  

if st.button('Get All Standard Data Frames'):
    csv_file_paths = get_file_paths_from_files_in_folder(export_path, ['.csv'])
    for file_path in csv_file_paths:
        full_file_path = f"{export_path}{file_path}"
        title_row_number = get_title_row_number(full_file_path, title_row_identifer='Name,Type', within_n_rows=3)-1
        df = pd.read_csv(full_file_path, header=title_row_number)
        st.subheader(file_path, divider="red")
        st.write(df.head(5))


if st.button('See sample data for distinct titles'):
    sorted_title_list = get_distinct_list_of_titles(export_path)
    dataframes_dict = get_dict__of_dataframes_of_all_csvs_in_folder(export_path)   
    
    # st.write(dataframes_dict)

    master_list = []

    # Iterate over each field
    for title in sorted_title_list:
        title_sample_data = []
        dataframe_names = dataframes_dict.keys()
        print(title)
        for dataframe_name in dataframe_names:
            df = dataframes_dict[dataframe_name]
            # print(df.columns)
            if title in df.columns:
                value_list = extract_field_data(df, title,1100)
                if value_list and len(value_list) > 0:
                    if len(value_list) == 1 and value_list[0].capitalize != 'NAN':
                        g=1
                    else:
                        title_sample_data.extend(value_list)
        unique_data = list(set(title_sample_data))
        master_list.append([title, unique_data])
    st.write(master_list)

    # for title, title_sample_data in master_list:
    #     st.subheader(title, divider="red")
    #     for dataframe_name, sample_data in title_sample_data:
    #         st.write(dataframe_name)
    #         st.write(sample_data)
if st.button('Consolidate CSVs'):
    csv_file_paths = get_file_paths_from_files_in_folder(export_path, ['.csv'])
    col_map = {
            'Required Black': 'Performance-Related',
            'Blue Trading': 'Trding-Related',
            'Default Value': 'Default Value',
            'Field Format': 'Definition2',
            'IC Definition': 'Definition',
            'is PII?': 'PII',
            'Name': 'DatabaseFieldName',
            'Nullable?': 'Nullable',
            'Orange Reporting': 'Reporting-Related',
            'Project Required': 'Project Required',
            'Required Field': 'Mandatory',
            'Table': 'Table'
        }
    consolidate_CSVs_to_one_file(column_map=col_map, csv_file_path_list=csv_file_paths)

    

