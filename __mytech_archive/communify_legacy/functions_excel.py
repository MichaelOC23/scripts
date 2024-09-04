import openpyxl





def read_excel_data(workbook_path, range_string):
    # Load the workbook
    workbook = openpyxl.load_workbook(workbook_path)

    if '!' in range_string:
        # Extract sheet name and range; strip single quotes from sheet name
        parts = range_string.split('!')
        sheet_name = parts[0].strip("'")
        column_range = parts[1]
    else:
        # Default to the first sheet
        sheet_name = workbook.sheetnames[0]
        column_range = range_string

    # Access the specified sheet
    sheet = workbook[sheet_name]

    # Read data from the specified column range
    data = []
    if ':' in column_range:  # If a full column/row range is specified
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
            data.append([cell.value for cell in row if cell.column_letter in column_range])
    else:  # If a single column/row or a single cell is specified
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                if cell.coordinate == column_range:
                    data.append(cell.value)
                    break

    return data

# Example usage
# data = read_excel_data('path_to_your_workbook.xlsx', "'consolidated copy 2'!E:E")


# Example usage
# data = read_excel_data('path_to_your_workbook.xlsx', "'consolidated copy 2'!E:E")


# Example usage
# data = read_excel_data('path_to_your_workbook.xlsx', "'consolidated copy 2'!E:E")
