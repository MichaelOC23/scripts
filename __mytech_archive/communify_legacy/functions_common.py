import streamlit as st  
import pandas as pd
import os
from streamlit_extras.row import row
import openpyxl



ROW_HEIGHT=72
appmode="live"
logopath="assets/logo.png"
oclogowidth=100
MARKDOWN_EXCEL_PATH = "assets/MarkdownExample.xlsx"



def get_current_user():
    return "michasmi"

def configure_streamlit_page(page_file_path=__file__, Title="Communify Product Horizons"):
    # Create session_state variables for the first two columns of every row in config.txt
    st.session_state['page_file_path'] = page_file_path
    
    
        
    # Create a settings dictionary in session_state if it doesn't, these are for theming and other settings
    if not "settings" in st.session_state:
        st.session_state['settings'] = {                
                "divider-color": "gray",
        }
        
    st.set_page_config(
            page_title="Communify Product Horizons",
            page_icon=":earth_americas:",
            layout="wide",
            initial_sidebar_state="collapsed",
            menu_items={
                'Get Help': 'mailto:michael@justbuildit.com',
                'Report a bug': 'mailto:michael@justbuildit.com',
                #'About': "# This is a header. This is an *extremely* cool app!"
                })
        
    # if "appmode" not in st.session_state:
    #     st.session_state['appmode'] = "None"
    
    # if st.session_state['appmode'] == "debug":
    #     with st.sidebar:
    #         st.markdown(f":orange[**Debug Mode:**] ON")
    #         st.markdown(f":blue[**File Name:**] {__file__.split('/')[-1]}""")
    #     session_col, config_col = st.columns([1.5,1])
    #     with session_col:
    #         with st.expander("Session State", expanded=False):
    #             st.write(st.session_state)
    #     with config_col:
    #         with st.expander("Config File", expanded=False):
    #             st.write(full_df)
                
                    
    ##############################
    ##   Standard Page Header   ##
    ##############################
    # Create a row with three columns
    headercol1, headercol2, headercol3, headercol4 = st.columns([.5 ,6, 1,1 ])
    # # Place the logo, title, and search button containers in the three columns
    # logo = headercol1.container(height=120, border=True)
    # title_text = headercol2.container(height=120, border=True)
    # search_button = headercol3.container(height=120, border=True)
    # Add the logo, title, and search button to the page
    headercol1.image("assets/OCLogo.png", width=45)        
    headercol2.subheader(Title)     
    # headercol3.button("Search", key="Search", use_container_width=True)
    headercol4.button("Sign Out", key="Logout", use_container_width=True)    

    st.markdown("<hr style='margin: 0.5em 0; border-top: 5px solid #ABABAB;'>", unsafe_allow_html=True)

def display_markdown_helper():
    def get_markdown_examples():
        # Load the workbook and select the first sheet
        workbook = openpyxl.load_workbook(MARKDOWN_EXCEL_PATH)
        first_sheet = workbook.worksheets[0]

        # Read the headers from the first row
        headers = [first_sheet.cell(row=1, column=j).value for j in range(1, 4)]

        # Read the rest of the rows
        data = []
        for i in range(2, first_sheet.max_row + 1):
            row_data = {}
            for j in range(1, 4):
                row_data[headers[j-1]] = first_sheet.cell(row=i, column=j).value
            data.append(row_data)

        return data
    Markdown_list = get_markdown_examples()
    
    for example in Markdown_list:
            
            #Prefix and suffix for the markdown to display the code
            Markdown_Prefix = "~~~   \n"
            Markdown_Suffix = "   \n~~~"
            
            #Create a container for the markdown and the title
            Markdown_Container = st.container(border=True,)
            Title_Nested_Container = Markdown_Container.container(height=45, border=False,)
            Title_Nested_Container.markdown(f"##### :violet[{example['Title']}]")
            Title_Nested_Container.markdown("<hr style='margin: 0.1em 0; border-top: 1px solid #bbb;'>", unsafe_allow_html=True)
            
            #Notes
            Markdown_Container.markdown(f"{example['Notes']}")    
            
            #Create columns for the markdown and the result
            b1, b2 = Markdown_Container.columns([1,1])
            
            #Add the ~~~ to the beginning and end of the markdown
            Markdown_Complete = f"{Markdown_Prefix}{example['Markdown']}{Markdown_Suffix}"
            
            #Markdown Code
            b1.markdown(":violet[Markdown Code]")
            b1.markdown(Markdown_Complete)
            
            #Result
            b2.markdown(":violet[Result]")
            b2.markdown(example['Markdown'], unsafe_allow_html=True)
            
if __name__ == "__main__":    
    configure_streamlit_page()
    display_markdown_helper()


    
    
        
        
        
    
    
    